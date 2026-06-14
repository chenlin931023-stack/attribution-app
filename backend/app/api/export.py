"""GET /api/export/{id} — 导出Excel大表"""
import os
import json
import io
from fastapi import APIRouter, HTTPException
from fastapi.responses import StreamingResponse
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

router = APIRouter(prefix="/api", tags=["export"])

RESULTS_DIR = os.path.expanduser("~/AttributionApp/results")


@router.get("/export/{task_id}")
def export_excel(task_id: str):
    result_path = os.path.join(RESULTS_DIR, f"{task_id}.json")
    if not os.path.exists(result_path):
        raise HTTPException(404, "结果不存在")

    with open(result_path, 'r', encoding='utf-8') as f:
        data = json.load(f)

    wb = Workbook()
    ws = wb.active
    ws.title = '业绩归因明细'

    pnav  = data['product_nav']
    assets = data['assets']
    fees   = data['fees']
    meta   = data['metadata']
    recon  = data.get('reconcile', {})

    DAYS = meta['days']
    date_s = meta['date_start']
    date_e = meta['date_end']
    PROD = meta['prod']
    prod_name = meta.get('prod_name', PROD)
    nav_init = pnav['nav_init']
    nav_end  = pnav['nav_end']
    ret_abs  = pnav['ret_abs']
    ret_ann  = pnav.get('ret_ann_vs_avg', 0) or 0
    nav_avg_wan = pnav.get('nav_avg_wan', 0) or 0

    THIN   = Side(style='thin', color='BFBFBF')
    MEDIUM = Side(style='medium', color='595959')
    B_THIN = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

    C_HDR1 = PatternFill('solid', fgColor='1F4E79')
    C_HDR2 = PatternFill('solid', fgColor='2E75B6')
    C_CAT = {
        '债券类资产管理计划': PatternFill('solid', fgColor='DEEAF1'),
        '公司债券': PatternFill('solid', fgColor='DEEAF1'),
        '货币存款': PatternFill('solid', fgColor='E2EFDA'),
        '股票类公募基金': PatternFill('solid', fgColor='FFF2CC'),
        '混合类公募基金': PatternFill('solid', fgColor='FFF2CC'),
        '产品费用': PatternFill('solid', fgColor='FCE4D6'),
    }
    C_TOTAL = PatternFill('solid', fgColor='D6DCE4')

    NC = 14
    # Title
    ws.merge_cells(f'A1:{get_column_letter(NC)}1')
    ws['A1'] = f'{PROD} {prod_name}  业绩归因明细  {date_s} ~ {date_e}（{DAYS}天）  年化: {ret_ann:.4f}%  日均规模: {nav_avg_wan:.0f}万'
    ws['A1'].fill = C_HDR1
    ws['A1'].font = Font(bold=True, color='FFFFFF', name='微软雅黑', size=11)
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 22

    # Summary
    ws.merge_cells(f'A2:{get_column_letter(NC)}2')
    ws['A2'] = f'期初NAV: {nav_init/1e8:.4f}亿 → 期末NAV: {nav_end/1e8:.4f}亿 | 净值法总收益: {pnav["total_return"]/1e4:.2f}万 | 校验差: {recon.get("gap_wan",0):.2f}万'
    ws['A2'].fill = C_HDR2
    ws['A2'].font = Font(color='FFFFFF', name='微软雅黑', size=9)
    ws['A2'].alignment = Alignment(horizontal='left', vertical='center', indent=1)
    ws.row_dimensions[2].height = 16

    # Headers
    headers = ['资产名称','资产类别','期初市值(万)','期末市值(万)','买卖收益(万)','公允变动(万)',
               '利息收入(万)','本期总收益(万)','对产品NAV贡献(%)','年化贡献(%)','HPR(%)','HPR年化(%)',
               '持仓年化(%)','备注']
    keys = ['资产名称','资产类别','期初市值(万)','期末市值(万)','本期买卖收益(万)','本期公允变动(万)',
            '本期利息收入(万)','本期总收益(万)','对产品NAV贡献率(%)','对产品年化贡献(%)','HPR(%)',
            'HPR年化(%)','平均持仓年化(%)','尾仓标注']
    for ci, h in enumerate(headers, 1):
        c = ws.cell(row=4, column=ci, value=h)
        c.fill = PatternFill('solid', fgColor='D6E4F0')
        c.font = Font(bold=True, name='微软雅黑', size=9)
        c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        c.border = B_THIN
    ws.row_dimensions[4].height = 30

    # Data rows
    rows = []
    def big_cat(cat):
        if cat == '托管账户': return '货币存款'
        return cat

    for a in assets:
        a['bc'] = big_cat(a.get('资产类别',''))
        rows.append(a)
    for f in fees:
        rows.append({'资产名称': f.get('费用类别',''), '资产类别':'产品费用', 'bc':'产品费用',
                     '期初市值(万)':0,'期末市值(万)':0,'本期买卖收益(万)':0,'本期公允变动(万)':0,
                     '本期利息收入(万)':0,'本期总收益(万)':f.get('本期费用(万)',0),
                     '对产品NAV贡献率(%)':f.get('对产品贡献率(%)',0),
                     '对产品年化贡献(%)':f.get('费用年化拖累_日均(%)',0),
                     'HPR(%)':None,'HPR年化(%)':None,'平均持仓年化(%)':None,
                     '尾仓标注': f'期间实付{f.get("期间实付(万)",0):.2f}万'})

    rows.sort(key=lambda r: (-(r.get('本期总收益(万)',0) or 0)))

    prev_cat = None
    for dr in rows:
        wr = ws.max_row + 1
        cat = dr.get('bc', dr.get('资产类别',''))
        is_total = (cat == '合计')
        is_new_cat = (cat != prev_cat)
        row_fill = C_CAT.get(cat, PatternFill('solid', fgColor='F2F2F2'))

        for ci, key in enumerate(keys, 1):
            val = dr.get(key, '')
            if val is None: val = ''
            c = ws.cell(row=wr, column=ci, value=val)
            c.fill = row_fill
            c.font = Font(bold=is_total, name='微软雅黑', size=9,
                          color='C00000' if isinstance(val, (int,float)) and val < -0.001 else '1F1F1F')
            c.alignment = Alignment(horizontal='right' if ci > 2 else 'left', vertical='center')
            c.border = Border(left=THIN, right=THIN, bottom=THIN, top=MEDIUM if is_new_cat else THIN)
        prev_cat = cat
        ws.row_dimensions[wr].height = 15

    col_widths = [30, 14, 11, 11, 11, 11, 11, 12, 11, 10, 9, 10, 10, 28]
    for ci, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(ci)].width = w
    ws.freeze_panes = 'C5'

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=attribution_{task_id}.xlsx"}
    )
